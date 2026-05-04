import openpyxl
from playwright.sync_api import sync_playwright
import time

# Load Excel file
wb = openpyxl.load_workbook('It23578500_Assignment_1_Test_Cases_.xlsx')
ws = wb['Test cases']

# Extract inputs
inputs = []
for row in range(2, ws.max_row + 1):
    input_val = ws.cell(row=row, column=3).value
    if input_val:
        inputs.append((row, str(input_val).strip()))

print(f"Found {len(inputs)} inputs to test")

# Run each input through the transliteration system
with sync_playwright() as p:
    browser = p.chromium.launch(headless=False)
    page = browser.new_page()
    
    try:
        # Open the transliteration website
        page.goto("https://www.pixelssuite.com/chat-translator", wait_until="domcontentloaded")
        page.wait_for_load_state("networkidle", timeout=30000)
        page.wait_for_selector("textarea", timeout=30000)
        print("Website loaded successfully")
        
        # Find the input and output textareas
        input_locator = page.locator('textarea[placeholder*="English"]').first
        output_locator = page.locator('textarea[placeholder*="Sinhala"]').first
        action_locator = page.get_by_role("button", name="Transliterate").first
        
        # Test each input
        previous_output = ""
        for row_num, input_text in inputs:
            print(f"Testing Row {row_num}: {input_text}")
            
            try:
                # Clear input and type new text
                input_locator.click()
                page.keyboard.press("Control+A")
                page.keyboard.press("Backspace")
                input_locator.fill(input_text)
                
                # Small delay to ensure input is registered
                page.wait_for_timeout(500)
                
                # Click transliterate button
                action_locator.click()
                
                # Additional delay after clicking
                page.wait_for_timeout(1000)
                
                # Wait for the output to be populated and different from previous
                max_attempts = 5
                actual_output = ""
                
                for attempt in range(max_attempts):
                    # Wait for output to be populated
                    try:
                        page.wait_for_function(
                            """() => {
                                const outputElement = document.querySelector('textarea[placeholder*="Sinhala"]');
                                if (!outputElement) return false;
                                const value = outputElement.value || outputElement.innerText || outputElement.textContent;
                                return value && value.trim() !== '';
                            }""",
                            timeout=8000
                        )
                    except:
                        print(f"  -> Timeout waiting for output on attempt {attempt + 1}")
                        continue
                    
                    # Wait a bit more for stability
                    page.wait_for_timeout(1500)
                    
                    # Get current output
                    try:
                        current_output = output_locator.input_value()
                        if not current_output:
                            current_output = output_locator.inner_text()
                        if not current_output:
                            current_output = output_locator.text_content()
                        
                        current_output = str(current_output).strip() if current_output else ""
                        
                        # Check if this is a new output (different from previous)
                        if current_output and current_output != previous_output:
                            actual_output = current_output
                            break
                        elif current_output and attempt == max_attempts - 1:
                            # Last attempt, use whatever we got
                            actual_output = current_output
                            break
                        else:
                            # Wait and try again
                            page.wait_for_timeout(2000)
                            
                    except Exception as e:
                        print(f"  -> Error getting output on attempt {attempt + 1}: {e}")
                        page.wait_for_timeout(2000)
                        continue
                
                # Update previous output for next iteration
                previous_output = actual_output
                
                # Get expected output from Excel
                expected_output = ws.cell(row=row_num, column=4).value
                expected_output = str(expected_output).strip() if expected_output else ""
                
                # Compare expected vs actual output and set status
                if expected_output and actual_output:
                    status = "PASS" if actual_output == expected_output else "FAIL"
                elif actual_output:
                    status = "FAIL"  # Has output but no expected to compare
                else:
                    status = "FAIL"  # No output
                
                # Update Excel
                ws.cell(row=row_num, column=5).value = actual_output  # Actual output column
                ws.cell(row=row_num, column=6).value = status  # Status column
                
                print(f"  -> Output: {actual_output}")
                
            except Exception as e:
                print(f"  -> Error: {e}")
                ws.cell(row=row_num, column=6).value = "ERROR"
            
            # Small delay between tests
            time.sleep(1)
        
        # Save results
        wb.save('It23578500_Assignment_1_Test_Cases_.xlsx')
        print("Results saved to Excel file")
        
    except Exception as e:
        print(f"Error loading website: {e}")
    
    browser.close()

print("Test execution completed")
